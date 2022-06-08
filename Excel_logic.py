import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class Cell:
    def __init__(self, value):
        self.value = value
        self.rowspan = 1
        self.colspan = 1


def coord_to_index(cells):
    indexs = []

    temp = cells[0]
    for j in range(1, len(cells)):
        if cells[j] == ':':
            continue
        if str(cells[j-1]).isalpha() != str(cells[j]).isalpha():
            indexs.append(temp)
            temp = cells[j]
        else:
            temp += cells[j]
    indexs.append(temp)
    for i in range(len(indexs)):
        if str(indexs[i]).isalpha():
            indexs[i] = column_index_from_string(indexs[i])
        else:
            indexs[i] = int(indexs[i])
    return indexs


class Excel:
    def __init__(self, filename):
        self.wb = openpyxl.load_workbook(filename, data_only=True)

    def get_data(self):
        self.wb.active = 0
        self.sheet = self.wb.active


        self.last_row = self.sheet.max_row
        self.last_col = self.sheet.max_column

        self.table = []
        for i in range(0, self.last_row):
            temp_row = []
            for j in range(0, self.last_col):
                if self.sheet.cell(i+1, j+1).value == None:
                    temp_row.append(Cell(''))
                else:
                    temp_row.append(Cell(self.sheet.cell(i+1, j+1).value))
            self.table.append(temp_row)


        merged_cells = self.sheet.merged_cells.ranges


        for i in range(len(merged_cells)):
            merged_cells[i] = coord_to_index(list(str(merged_cells[i])))


        for cell in range(len(merged_cells)):
            block = merged_cells[cell]

            for i in range(block[1]-1, block[3]):
                for j in range(block[0]-1, block[2]):
                    self.table[i][j].rowspan = 0
                    self.table[i][j].colspan = 0
            self.table[block[1]-1][block[0]-1].rowspan = block[3] - block[1] + 1
            self.table[block[1]-1][block[0]-1].colspan = block[2] - block[0] + 1


        """for i in range(len(self.table)):
            temp = ''
            for j in range(len(self.table[i])):
                temp += str(self.table[i][j].value) + ' '
            print(temp)

        for i in range(len(self.table)):
            temp = ''
            for j in range(len(self.table[i])):
                temp += '[' + str(self.table[i][j].rowspan) + ' ' + str(self.table[i][j].colspan) + '] '
            print(temp)"""


"""exp = Excel("sample.xlsx")
exp.get_data()"""

